"""File handling utilities for poems data."""

from typing import List, Optional
import os

from src.models.poem import Poem


class FileHandler:
    """
    Handles reading and writing poem data to files.
    """
    
    @staticmethod
    def load_poems_from_file(filepath: str) -> Optional[List[Poem]]:
        """
        Load poems from a text file.
        
        Expected format: "Título | Autor | Año | Género" (one per line)
        Lines starting with # are ignored (comments)
        
        Args:
            filepath: Path to the input file
            
        Returns:
            List of Poem objects or None if file doesn't exist
        """
        if not os.path.exists(filepath):
            return None
        
        poems = []
        try:
            with open(filepath, 'r', encoding='utf-8') as f:
                for idx, line in enumerate(f, 1):
                    line = line.strip()
                    
                    # Skip empty lines and comments
                    if not line or line.startswith('#'):
                        continue
                    
                    poem = Poem.create_from_text(idx, line)
                    if poem:
                        poems.append(poem)
            
            return poems if poems else None
            
        except Exception as e:
            print(f"Error loading poems from file: {e}")
            return None
    
    @staticmethod
    def save_to_excel(poems: List[Poem], filepath: str) -> bool:
        """
        Save poems to Excel file.
        
        Args:
            poems: List of Poem objects
            filepath: Output filepath
            
        Returns:
            True if successful, False otherwise
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            
            # Create workbook and sheet
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Poemas Dominicanos"
            
            # Headers
            headers = list(poems[0].to_dict().keys())
            ws.append(headers)
            
            # Style headers
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Add data
            for poem in poems:
                row_data = list(poem.to_dict().values())
                ws.append(row_data)
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 80)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Color-code by availability
            for row_idx, poem in enumerate(poems, start=2):
                availability_cell = ws.cell(row=row_idx, column=len(headers))
                
                if poem.disponibilidad == "ENCONTRADO":
                    availability_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    availability_cell.font = Font(color="006100")
                elif poem.disponibilidad == "PARCIAL":
                    availability_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                    availability_cell.font = Font(color="9C5700")
                else:
                    availability_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    availability_cell.font = Font(color="9C0006")
            
            # Save file
            wb.save(filepath)
            return True
            
        except Exception as e:
            print(f"Error saving to Excel: {e}")
            return False
    
    @staticmethod
    def save_to_csv(poems: List[Poem], filepath: str) -> bool:
        """
        Save poems to CSV file.
        
        Args:
            poems: List of Poem objects
            filepath: Output filepath
            
        Returns:
            True if successful, False otherwise
        """
        try:
            import csv
            
            with open(filepath, 'w', newline='', encoding='utf-8') as f:
                if not poems:
                    return False
                
                # Get headers from first poem
                headers = list(poems[0].to_dict().keys())
                
                writer = csv.DictWriter(f, fieldnames=headers)
                writer.writeheader()
                
                for poem in poems:
                    writer.writerow(poem.to_dict())
            
            return True
            
        except Exception as e:
            print(f"Error saving to CSV: {e}")
            return False
